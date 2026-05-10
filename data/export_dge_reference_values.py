#!/usr/bin/env python3

from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
INPUT_XLSX = ROOT / "DGE-Referenzwerte.xlsx"
OUTPUT_JSON = ROOT / "dge-referenzwerte.json"
SOURCE_SHEET = "Referenzwerte"


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def build_payload(xlsx_path: Path) -> dict:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = workbook[SOURCE_SHEET]

    headers = [sheet.cell(1, c).value for c in range(1, 8)]
    col = {name: idx + 1 for idx, name in enumerate(headers)}

    entries = []
    for row in range(2, sheet.max_row + 1):
        population_group = cell_text(sheet.cell(row, col["Bevölkerungsgruppe"]).value)
        sex = cell_text(sheet.cell(row, col["Geschlecht"]).value)
        nutrient = cell_text(sheet.cell(row, col["Nährstoff"]).value)
        reference_value = cell_text(sheet.cell(row, col["Referenzwert"]).value)
        unit = cell_text(sheet.cell(row, col["Einheit"]).value)

        if not nutrient and not reference_value and not unit:
            continue

        category = cell_text(sheet.cell(row, col["Kategorie"]).value)
        remark = cell_text(sheet.cell(row, col["Bemerkung"]).value)
        label = f"{nutrient} - {reference_value}" + (f" {unit}" if unit else "")

        entries.append(
            {
                "populationGroup": population_group,
                "sex": sex,
                "nutrient": nutrient,
                "referenceValue": reference_value,
                "unit": unit,
                "label": label,
                "category": category,
                "remark": remark,
            }
        )

    return {
        "version": 1,
        "source": {
            "file": xlsx_path.name,
            "sheet": SOURCE_SHEET,
        },
        "dimensions": {
            "populationGroups": sorted({entry["populationGroup"] for entry in entries if entry["populationGroup"]}),
            "sexes": sorted({entry["sex"] for entry in entries if entry["sex"]}),
        },
        "entries": entries,
    }


def main():
    payload = build_payload(INPUT_XLSX)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
